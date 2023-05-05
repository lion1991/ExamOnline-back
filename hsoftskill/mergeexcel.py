import os
import xlrd2
import xlsxwriter
import glob
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Side, Border
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
class MakeExcel():
    # exam_user = ['葛秀峰', '李健卓', '薛俞同', '王金龙', '殷晓宇', '徐嘉霖', '刘哲', '柯鹏飞', '刘树国', '季恩光', '张成全',
    #              '王彦淇', '汪鑫泽', '邵成', '赵宇轩', '郭大利', '王景东', '王超', '关涛', '王旭明', '王景鑫', '霍锳林', '高智勇',
    #              '张德龙', '姜超', '董洋', '李幸龙', '刘泽鹏', '冷步凡', '刘博', '焉海龙', '姜伟宁', '杜伟', '陈懋峰', '赵鑫宇',
    #              '田照明', '吕俊鹏', '吕云鹏', '杨枭', '朱玉', '梅杰', '于利强', '宋云鹏', '张士文', '汪浩炀', '夏思豪']
    name_dict =  {1 : (1,1,'葛秀峰'), 2 : (2,1,'李健卓'), 3 : (3,1,'薛俞同'), 4 : (4,1,'王金龙'), 5 : (5,1,'殷晓宇'),
                  6 : (6,1,'徐嘉霖'), 7 : (7,1,'刘哲'), 8 : (8,2,'柯鹏飞'),
                  9 : (9,2,'刘树国'), 10 : (10,2,'季恩光'), 11 : (11,2,'张成全'), 12 : (12,2,'王彦淇'), 13 : (13,2,'汪鑫泽'),
                  14 : (14,2,'邵成'), 15 : (15,2,'赵宇轩'),
                  16 : (16,3,'郭大利'), 17 : (17,3,'王景东'), 18 : (18,3,'王超'), 19 : (19,3,'关涛'), 20 : (20,3,'王旭明'),
                  21 : (21,3,'王景鑫'), 22 : (22,3,'霍锳林'),
                  23 : (23,4,'高智勇'), 24 : (24,4,'张德龙'), 25 : (25,4,'姜超'), 26 : (26,4,'董洋'), 27 : (27,4,'李幸龙'),
                  28 : (28,4,'刘泽鹏'), 29 : (29,4,'冷步凡'), 30 : (30,4,'刘博'),
                  31 : (31,5,'焉海龙'), 32 : (32,5,'姜伟宁'), 33 : (33,5,'杜伟'), 34 : (34,5,'陈懋峰'), 35 : (36,5,'赵鑫宇'),
                  36 : (36,5,'田照明'), 37 : (37,5,'吕俊鹏'),
                  38 : (38,5,'吕云鹏'), 39 : (39,6,'杨枭'), 40 : (40,6,'朱玉'), 41 : (41,6,'梅杰'), 42 : (42,6,'于利强'),
                  43 : (43,6,'宋云鹏'), 44 : (44,6,'张士文'), 45 : (45,6,'汪浩炀'),46 : (46,6,'夏思豪')}
    biao_tou_total1_1 = "NULL"
    biao_tou_total1_2 = "NULL"
    biao_tou_total2_1 = "NULL"
    biao_tou_total2_2 = "NULL"
    biao_tou_total3_1 = "NULL"
    biao_tou_total3_2 = "NULL"
    wei_zhi = "NULL"
    # 保存第一个sheet数据
    all_data1 = []
    # 保存第二个sheet数据
    all_data2 = []
    # 保存第三个sheet数据
    all_data3 = []

    # 用于保存合并的所有行的数据
    # 获取要合并的所有exce表格
    def get_exce(self, filepath):
        # wei_zhi = input("请输入Excel文件所在的目录：")
        self.wei_zhi = (filepath)
        all_exce = glob.glob(self.wei_zhi + "*.xlsx")
        print("该目录下有" + str(len(all_exce)) + "个excel文件：")
        if (len(all_exce) == 0):
            return 0
        else:
            for i in range(len(all_exce)):
                print(all_exce[i])
            return all_exce

    # 打开Exce文件
    def open_exce(self, name):
        fh = xlrd2.open_workbook(name)
        return fh

    # 获取exce文件下的所有sheet名
    def get_sheet(self, fh):
        sheets = fh.sheets()
        #print(sheets)
        return sheets

    # 获取sheet下有多少行数据
    def get_sheetrow_num(self, sheet):
        return sheet.nrows

    # 获取sheet下的数据
    def get_sheet_data(self, sheet, row):
        for i in range(row):
            if (i == 0):
                self.biao_tou_total1_1 = sheet.row_values(i)
                continue
            # if (i == 1):
            #     self.biao_tou_total1_2 = sheet.row_values(i)
            #     continue
            values = sheet.row_values(i)
            print(values)
            self.all_data1.append(values)

        return self.all_data1

    def get_sheet_data2(self, sheet, row):
        for i in range(row):
            if (i == 0):
                self.biao_tou_total2_1 = sheet.row_values(i)
                continue
            if (i == 1):
                self.biao_tou_total2_2 = sheet.row_values(i)
                continue

            values = sheet.row_values(i)
            # print(values)
            self.all_data2.append(values)

        return self.all_data2

    def get_sheet_data3(self, sheet, row):
        for i in range(row):
            if (i == 0):
                self.biao_tou_total3_1 = sheet.row_values(i)
                continue
            if (i == 1):
                self.biao_tou_total3_2 = sheet.row_values(i)
                continue

            values = sheet.row_values(i)
            # print(values)
            self.all_data3.append(values)

        return self.all_data3

    def startwork(self, filepath):
        #数据初始化
        self.biao_tou_total1_1 = "NULL"
        self.biao_tou_total1_2 = "NULL"
        self.biao_tou_total2_1 = "NULL"
        self.biao_tou_total2_2 = "NULL"
        self.biao_tou_total3_1 = "NULL"
        self.biao_tou_total3_2 = "NULL"
        self.wei_zhi = "NULL"
        # 保存第一个sheet数据
        self.all_data1 = []
        # 保存第二个sheet数据
        self.all_data2 = []
        # 保存第三个sheet数据
        self.all_data3 = []


        all_exce = self.get_exce(filepath)
        sheets_count = 0  #存放考核分类数量
        examtype = ''
        # 得到要合并的所有exce表格数据
        if (all_exce == 0):
            print("该目录下无.xlsx文件！请检查您输入的目录是否有误！")
            os.system('pause')
            exit()

        # 下面开始文件数据的获取
        for exce in all_exce:
            fh = self.open_exce(exce)
            # 打开文件
            sheets = self.get_sheet(fh)
            # examname = str(sheets[1])
            # 打开的sheet名
            examtype = '组员评分'
            # if "服务器" in examname:
            #     examtype = '服务器题'
            # else:
            #     examtype = '数通题'
            # print(sheet_name)
            sheets_count = len(sheets)

            # 获取文件下的sheet数量

            for sheet in range(len(sheets)):
                # print(sheet)
                row = self.get_sheetrow_num(sheets[sheet])
                # print(row)
                # 获取原表中一个sheet下的所有的数据的行数
                while sheet == 0:
                    sheet1 = self.get_sheet_data(sheets[sheet], row)
                    break
                # try:
                #     while sheet == 1:
                #         sheet2 = self.get_sheet_data2(sheets[sheet], row)
                #         break
                # except:
                #     print("sheet2不存在")
                # try:
                #     while sheet == 2:
                #         sheet3 = self.get_sheet_data3(sheets[sheet], row)
                #         break
                # except:
                #     print("sheet3不存在")
                # 获取一个sheet下的所有行的数据
        try:
            # self.all_data1.insert(0, self.biao_tou_total1_2)
            self.all_data1.insert(0, self.biao_tou_total1_1)
        except:
            print("sheet1不存在")
        # try:
        #     self.all_data2.insert(0, self.biao_tou_total2_2)
        #     self.all_data2.insert(0, self.biao_tou_total2_1)
        # except:
        #     print("sheet2不存在")
        # try:
        #     if sheets_count > 2:
        #         self.all_data3.insert(0, self.biao_tou_total3_2)
        #         self.all_data3.insert(0, self.biao_tou_total3_1)
        # except:
        #     print("sheet3不存在")
        # 表头写入

        # 下面开始文件数据的写入
        new_exce = self.wei_zhi + "组员评分汇总.xlsx"
        # 新建的exce文件名字

        fh1 = xlsxwriter.Workbook(new_exce)
        # 新建一个exce表
        # print(sheets_count)
        try:
            new_sheet1 = fh1.add_worksheet()
            # new_sheet2 = fh1.add_worksheet()
            # if sheets_count > 2:
            #     new_sheet3 = fh1.add_worksheet()
        except:
            print("创建新表格sheet出错")
        # 新建一个sheet表
        # 写入sheet表数据
        try:
            for i in range(len(self.all_data1)):
                for j in range(len(self.all_data1[i])):
                    c = self.all_data1[i][j]
                    new_sheet1.write(i, j, c)
        except:
            print("sheet1不存在,合成表无数据写入。")
        # try:
        #     for i in range(len(self.all_data2)):
        #         for j in range(len(self.all_data2[i])):
        #             c = self.all_data2[i][j]
        #             new_sheet2.write(i, j, c)
        # except:
        #     print("sheet2不存在,合成表无数据写入。")
        # try:
        #     if sheets_count > 2:
        #         for i in range(len(self.all_data3)):
        #             for j in range(len(self.all_data3[i])):
        #                 c = self.all_data3[i][j]
        #                 new_sheet3.write(i, j, c)
        # except:
        #     print("sheet3不存在,合成表无数据写入。")
        fh1.close()
        # 关闭该exce表

        print("文件合并成功,请查看“" + self.wei_zhi + "”目录下的组员评分汇总.xlsx文件！")

        #找出未上传答题的人员加入表格最后

        # 打开 Excel 文件
        workbook = load_workbook(filename=self.wei_zhi + '组员评分汇总.xlsx')

        # 遍历工作簿中的所有工作表
        for sheet in workbook:
            # 获取工作表的名称和工作表对象
            sheet_name = sheet.title
            sheet_data = sheet.values

            # 跳过标题行
            # next(sheet_data)
            # next(sheet_data)
            # 遍历第二列第三行开始的所有单元格，并将其添加到集合中
            cell_values = set()
            for row in sheet_data:
                cell_value = row[2]
                if cell_value:
                    cell_values.add(cell_value)
            # 查找字典中没有的姓名
            for key, value in self.name_dict.items():
                if value[2] not in cell_values:
                    max_row = sheet.max_row
                    new_row = [value[0], f'第{value[1]}组', value[2]]
                    for i in range(4, sheet.max_column + 1):
                        new_row.append(0)
                    sheet.append(new_row)

            cell_values.clear()
        # 保存 Excel 文件
        workbook.save(self.wei_zhi + '组员评分汇总.xlsx')

        #开始修改格式

        workbook = load_workbook(filename=self.wei_zhi + '组员评分汇总.xlsx')
        # font = Font(name='宋体', size=18, bold=True)
        # sheetnames = workbook.sheetnames
        # mo_sheet = workbook[sheetnames[0]]
        # mo_sheet['A1'].font = font
        #修改边框
        border = Border(left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin'))
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for ws in workbook:
            for row in ws.rows:
                for cell in row:
                    # 如果单元格有数据，则为它添加边框
                    if cell.value or cell.value == 0:
                        cell.border = border
                        cell.alignment = alignment
        mo_sheet_one = workbook['Sheet1']
        mo_sheet_one.title = '组员评分'
        #     mo_sheet_one.title = '总成绩'
        # if sheets_count > 2:
        #     mo_sheet_one = workbook['Sheet1']
        #     mo_sheet_one.title = '总成绩'
        #     mo_sheet_two = workbook['Sheet2']
        #     mo_sheet_two.title = examtype
        #     mo_sheet_three = workbook['Sheet3']
        # else:
        #     mo_sheet_one = workbook['Sheet1']
        #     mo_sheet_one.title = '总成绩'
        #     mo_sheet_two = workbook['Sheet2']
        #     mo_sheet_two.title = examtype
        #设置指定列宽

        #获取表1总列数
        column_count = mo_sheet_one.max_column

        #字母自增
        # _chr = 'A'
        # if _chr == 'Z':
        #     print('A')
        # else:
        #     print(chr(ord(_chr) + 1))
        #对表1进行列宽设置
        _chr = 'A'
        for i in range(column_count):

            if i == 0:
                mo_sheet_one.column_dimensions[chr(ord(_chr))].width = 5
                _chr = chr(ord(_chr) + 1)
                continue
            if i == 1:
                mo_sheet_one.column_dimensions[chr(ord(_chr))].width = 10
                _chr = chr(ord(_chr) + 1)
                continue
            if i == 2:
                mo_sheet_one.column_dimensions[chr(ord(_chr))].width = 10
                _chr = chr(ord(_chr) + 1)
                continue
            else:
                # print(chr(ord(_chr)))
                mo_sheet_one.column_dimensions[chr(ord(_chr))].width = 20
                _chr = chr(ord(_chr) + 1)

        # #对表2进行列宽设置
        # _chr = 'A'
        # #获取表2总列数
        # column_count2 = mo_sheet_two.max_column
        # for i in range(column_count2):
        #
        #     if i == 0:
        #         mo_sheet_two.column_dimensions[chr(ord(_chr))].width = 5
        #         _chr = chr(ord(_chr) + 1)
        #         continue
        #     if i == 1:
        #         mo_sheet_two.column_dimensions[chr(ord(_chr))].width = 10
        #         _chr = chr(ord(_chr) + 1)
        #         continue
        #     else:
        #         # print(chr(ord(_chr)))
        #         mo_sheet_two.column_dimensions[chr(ord(_chr))].width = 20
        #         _chr = chr(ord(_chr) + 1)
        #
        # #对表3进行列宽设置
        # if sheets_count > 2:
        #     _chr = 'A'
        #     #获取表3总列数
        #     column_count3 = mo_sheet_three.max_column
        #     for i in range(column_count3):
        #
        #         if i == 0:
        #             mo_sheet_three.column_dimensions[chr(ord(_chr))].width = 5
        #             _chr = chr(ord(_chr) + 1)
        #             continue
        #         if i == 1:
        #             mo_sheet_three.column_dimensions[chr(ord(_chr))].width = 10
        #             _chr = chr(ord(_chr) + 1)
        #             continue
        #         else:
        #             # print(chr(ord(_chr)))
        #             mo_sheet_three.column_dimensions[chr(ord(_chr))].width = 20
        #             _chr = chr(ord(_chr) + 1)


        #合并表1单元格
        # if sheets_count > 2:
        #     mo_sheet_one.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
        #     mo_sheet_one.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        #     mo_sheet_one.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        # else:
        #     mo_sheet_one.merge_cells(start_row=1, start_column=3, end_row=1, end_column=3)
        #     mo_sheet_one.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        #     mo_sheet_one.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        # 合并表2单元格
        # mo_sheet_two.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        # mo_sheet_two.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        # row_values = []
        # init_value = []
        # row_num = 0
        # increase_num = 0
        # for row in mo_sheet_two.iter_rows(min_col=1, max_col=mo_sheet_two.max_column, max_row=1):
        #     # row_values.append([cell.value for cell in row])
        #     for cell in row:
        #         if cell.value == '总分':
        #             init_value.append(increase_num)
        #         elif cell.value != None:
        #             increase_num += 1
        #             row_num += 1
        #         elif cell.value == None:
        #             increase_num += 1
        #             if row_num not in init_value:
        #                 init_value.append(row_num)
        #
        # mo_sheet_two.merge_cells(start_row=1, start_column=init_value[0], end_row=1, end_column=init_value[1])
        # 合并表3单元格
        # if sheets_count > 2:
        #     mo_sheet_three.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        #     mo_sheet_three.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        #     row_values = []
        #     init_value = []
        #     row_num = 0
        #     increase_num = 0
        #     for row in mo_sheet_three.iter_rows(min_col=1, max_col=mo_sheet_three.max_column, max_row=1):
        #         # row_values.append([cell.value for cell in row])
        #         for cell in row:
        #             if cell.value == '总分':
        #                 init_value.append(increase_num)
        #             elif cell.value != None:
        #                 increase_num += 1
        #                 row_num += 1
        #             elif cell.value == None:
        #                 increase_num += 1
        #                 if row_num not in init_value:
        #                     init_value.append(row_num)
        #
        #     mo_sheet_three.merge_cells(start_row=1, start_column=init_value[0], end_row=1, end_column=init_value[1])

        #定义临时列表存放数据
        data = []
        #开始按照序号从小到大排序第一个工作表
        for row in mo_sheet_one.iter_rows(min_row=2, values_only=True):
            data.append(row)
        data.sort(key=lambda x: x[0])
        row_num = 2
        for i, row in enumerate(data):
            for j, cell in enumerate(row):
                mo_sheet_one.cell(row=row_num, column=j + 1).value = cell
            row_num += 1
        data.clear()
        # 打印工作表的内容
        # for row in mo_sheet_one.values:
        #     print(row)
        #开始按照序号从小到大排序第二个工作表
        # for row in mo_sheet_two.iter_rows(min_row=3, values_only=True):
        #     data.append(row)
        # data.sort(key=lambda x: x[0])
        # for i, row in enumerate(data):
        #     for j, cell in enumerate(row):
        #         mo_sheet_two.cell(row=i + 3, column=j + 1).value = cell
        # data.clear()
        # 打印工作表的内容
        # for row in mo_sheet_two.values:
        #     print(row)

        #设置表头行高
        mo_sheet_one.row_dimensions[1].height = 56
        #设置数据行行高
        for row in range(2,mo_sheet_one.max_row + 1):
            mo_sheet_one.row_dimensions[row].height = 26
        # for row in range(1,mo_sheet_two.max_row + 1):
        #     if row == 2:
        #         mo_sheet_two.row_dimensions[row].height = 55
        #     else:
        #         mo_sheet_two.row_dimensions[row].height = 30
        # if sheets_count > 2:
        #     for row in range(1,mo_sheet_three.max_row + 1):
        #         if row == 2:
        #             mo_sheet_three.row_dimensions[row].height = 55
        #         else:
        #             mo_sheet_three.row_dimensions[row].height = 30

        workbook.save(self.wei_zhi + '组员评分汇总.xlsx')