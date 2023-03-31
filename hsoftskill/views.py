
from django.http import FileResponse, Http404, JsonResponse
from django.http import HttpResponse

import os
import random
import zipfile
import math

# Create your views here.
from rest_framework import mixins, viewsets, generics, status
from rest_framework.decorators import action
from rest_framework.generics import GenericAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.settings import api_settings
from rest_framework.utils import json
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from rest_framework_simplejwt.authentication import JWTAuthentication

from exam.views import CommonPagination
from hsoftskill.filevars import fileuserid, qj_userid, ljj_userid, wcl_userid, ybw_userid, wy_userid, \
    captainName, downloadlink, lc_userid ,filerootpath
from hsoftskill import serializers, robot, mergeexcel
from hsoftskill.models import Files, LimitFile, RandomFileId, CaptainFiles, ExamFiles, JudgeFiles, PersonalGradeFiles
#from tyadmin_api.auto_serializers import FilesListSerializer
#from tyadmin_api.auto_views import FilesViewSet
#from tyadmin_api.custom import XadminViewSet
#合并文件模块导入
# import os
# import xlrd2
# import xlsxwriter
# import glob
# from openpyxl import load_workbook
# from openpyxl.styles import Alignment
# from openpyxl.styles import Side, Border
# from openpyxl.styles import Font


class ExamFileView(generics.CreateAPIView):
    """
    上传考核题文件
    """
    authentication_classes = []
    permission_classes = []
    queryset = ExamFiles.objects.filter()
    serializer_class = serializers.ExamFileSerializer

    def create(self, request, *args, **kwargs):

        uploader = request.data.get('uploader')
        period = request.data.get('period')
        query = ExamFiles.objects.filter(period=period).values('uploader')[::1]
        uploadedid = []
        for index in query:
            uploadedid.append(index.get('uploader'))
        if int(uploader) in uploadedid:
            return Response(data={'msg': '处理失败，已经上传过', 'code': 405}, status=200)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        content = "考核题已经发布，请到网站下载。"
        is_send_all = False
        robot.send_ding_message(content,is_send_all)
        return Response(data={'data': serializer.data, 'code': 200}, status=status.HTTP_201_CREATED, headers=headers)


class FileView(generics.CreateAPIView):
    """
    上传组员考核文件
    """
    authentication_classes = []
    permission_classes = []
    queryset = Files.objects.filter()
    serializer_class = serializers.FileSerializer

    def create(self, request, *args, **kwargs):
        # print(request.data)
        # if request.data.get('period') == "":
        #     return Response(data={'code':401,'msg':"参数不全"})
        # 过滤文件名是否符合规范
        # namelist = ['张三']
        # filename = request.data.get('name').split('.')[0]
        # if filename in namelist:
        uploader = request.data.get('uploader')
        period = request.data.get('period')
        query = Files.objects.filter(period=period).values('uploader')[::1]
        uploadedid = []
        for index in query:
            uploadedid.append(index.get('uploader'))
        if int(uploader) in uploadedid:
            return Response(data={'msg': '处理失败，已经上传过', 'code': 405}, status=200)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(data={'data': serializer.data, 'code': 200}, status=status.HTTP_201_CREATED, headers=headers)
        # else:
        #     print(filename)
        #     return Response(data={'code': 402, 'msg': "文件命名不正确"})

class CaptainExamFileView(generics.CreateAPIView):
    """
    上传组长考核文件
    """
    authentication_classes = []
    permission_classes = []
    queryset = CaptainFiles.objects.filter()
    serializer_class = serializers.CaptainFileSerializer

    def create(self, request, *args, **kwargs):

        uploader = request.data.get('uploader')
        period = request.data.get('period')
        query = CaptainFiles.objects.filter(period=period).values('uploader')[::1]
        uploadedid = []
        for index in query:
            uploadedid.append(index.get('uploader'))
        if int(uploader) in uploadedid:
            return Response(data={'msg': '处理失败，已经上传过', 'code': 405}, status=200)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(data={'data': serializer.data, 'code': 200}, status=status.HTTP_201_CREATED, headers=headers)

class JudgeFileView(generics.CreateAPIView):
    """
    上传评分文件
    """
    authentication_classes = []
    permission_classes = []
    queryset = JudgeFiles.objects.filter()
    serializer_class = serializers.JudgeFileSerializer

    def create(self, request, *args, **kwargs):

        uploader = request.data.get('uploader')
        period = request.data.get('period')
        query = JudgeFiles.objects.filter(period=period).values('uploader')[::1]
        uploadedid = []
        for index in query:
            uploadedid.append(index.get('uploader'))
        if int(uploader) in uploadedid:
            return Response(data={'msg': '处理失败，已经上传过', 'code': 405}, status=200)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(data={'data': serializer.data, 'code': 200}, status=status.HTTP_201_CREATED, headers=headers)

class PersonalGradeFileView(generics.CreateAPIView):
    """
    上传个人成绩文件
    """
    authentication_classes = []
    permission_classes = []
    queryset = PersonalGradeFiles.objects.filter()
    serializer_class = serializers.PersonalGradeFileSerializer

    def create(self, request, *args, **kwargs):

        uploader = request.data.get('uploader')
        period = request.data.get('period')
        query = PersonalGradeFiles.objects.filter(period=period).values('uploader')[::1]
        uploadedid = []
        for index in query:
            uploadedid.append(index.get('uploader'))
        if int(uploader) in uploadedid:
            return Response(data={'msg': '处理失败，已经上传过', 'code': 405}, status=200)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(data={'data': serializer.data, 'code': 200}, status=status.HTTP_201_CREATED, headers=headers)

class ExamFileDownload(ModelViewSet):
    """
    考核题下载
    """
    authentication_classes = []
    permission_classes = []
    queryset = ExamFiles.objects.filter()
    serializer_class = serializers.ExamFileSerializer

    # 文件下载响应
    @action(methods=['get', 'post'], detail=True)
    def download(self, request, pk=None, *args, **kwargs):
        file_obj = self.get_object()
        print(file_obj)
        filename = str(file_obj)
        response = FileResponse(open(file_obj.file.path, 'rb'))
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="{}"'.format(filename.encode('utf-8').decode('ISO-8859-1'))
        return response


class FileDownload(ModelViewSet):
    """
    获取组员考核文件下载
    """
    authentication_classes = []
    permission_classes = []
    queryset = Files.objects.filter()
    serializer_class = serializers.FileSerializer

    # 文件下载响应

    def download_zip(self, request, filename, period):
        file_path = os.path.join(filerootpath + '/exam' + str(period) + '/' + filename)
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/zip")
                response['Content-Disposition'] = 'attachment;filename="{}"'.format(filename.encode('utf-8').decode('ISO-8859-1'))
                return response
        raise Http404

class CaptainFileDownload(ModelViewSet):
    """
    获取组长考核文件下载
    """
    authentication_classes = []
    permission_classes = []
    queryset = CaptainFiles.objects.filter()
    serializer_class = serializers.CaptainFileSerializer
    # 文件下载响应
    @action(methods=['get', 'post'], detail=True)
    def download(self, request, pk=None, *args, **kwargs):
        file_obj = self.get_object()
        print(file_obj)
        filename = str(file_obj)
        response = FileResponse(open(file_obj.file.path, 'rb'))
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="{}"'.format(filename.encode('utf-8').decode('ISO-8859-1'))
        return response

class JudgeFileDownload(ModelViewSet):
    """
    评分文件下载
    """
    authentication_classes = []
    permission_classes = []
    queryset = JudgeFiles.objects.filter()
    serializer_class = serializers.JudgeFileSerializer

    # 文件下载响应
    @action(methods=['get', 'post'], detail=True)
    def download(self, request, pk=None, *args, **kwargs):
        file_obj = self.get_object()
        print(file_obj)
        filename = str(file_obj)
        response = FileResponse(open(file_obj.file.path, 'rb'))
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="{}"'.format(filename.encode('utf-8').decode('ISO-8859-1'))
        return response

class PersonalGradeFileDownload(ModelViewSet):
    """
    个人成绩文件下载
    """
    authentication_classes = []
    permission_classes = []
    queryset = PersonalGradeFiles.objects.filter()
    serializer_class = serializers.PersonalGradeFileSerializer

    # 文件下载响应
    @action(methods=['get', 'post'], detail=True)
    def download(self, request, pk=None, *args, **kwargs):
        file_obj = self.get_object()
        print(file_obj)
        filename = str(file_obj)
        response = FileResponse(open(file_obj.file.path, 'rb'))
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="{}"'.format(filename.encode('utf-8').decode('ISO-8859-1'))
        return response

class UploadListViewSet(mixins.ListModelMixin, mixins.CreateModelMixin, viewsets.GenericViewSet):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    """showfile 查询上传列表"""
    # authentication_classes = []
    # permission_classes = []
    # 这里必须要定义一个默认的排序,否则会报错
    queryset = Files.objects.all().order_by('-upload_time')
    # 序列化
    serializer_class = serializers.FileSerializer
    # 分页
    pagination_class = CommonPagination

    # 重写queryset
    def get_queryset(self):
        # 学生ID
        student_id = self.request.query_params.get("student_id")
        if student_id:
            self.queryset = Files.objects.filter(uploader_id=student_id)
        return self.queryset

    # 重写返回
    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        result_data = response.data
        return Response(
            {
                "code": 200,
                "message": "查询成功",
                "response": result_data
            }
        )


class LimitPeriodListViewSet(mixins.ListModelMixin, mixins.CreateModelMixin, viewsets.GenericViewSet):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    """showlimitperiod，查询上传限制"""
    # authentication_classes = []
    # permission_classes = []
    # 这里必须要定义一个默认的排序,否则会报错
    queryset = LimitFile.objects.all()
    # 序列化
    serializer_class = serializers.LimitFileSerializer

    # 重写返回
    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        result_data = response.data
        return Response(
            {
                "code": 200,
                "message": "查询成功",
                "response": result_data
            }
        )



class FilePersion(APIView):
    # 返回查询的文件上传信息
    """
    生成各组长判分的组员，管理员每期结束后手动调用一次
    """
    authentication_classes = []
    permission_classes = []

    def get(self, request):
        req = request.GET
        period = req['period']

        # 获取上传目录下的所有文件名并随机排序
        upload_dir = filerootpath + "/exam" + str(period)
        file_names = os.listdir(upload_dir)
        random.shuffle(file_names)

        # 计算每组文件数量
        group_size = 6
        num_files = len(file_names)
        files_per_group = math.ceil(num_files / group_size)

        # 将文件名分组
        groups = [[] for _ in range(group_size)]
        for i, file_name in enumerate(file_names):
            group_index = i % group_size
            groups[group_index].append(file_name)

        # 将剩余的文件添加到最后一组中
        #if len(groups[-1]) < files_per_group:
        #    groups[-1].extend(groups.pop())

        # 逐个分组进行压缩
        download_links = []
        for i, group in enumerate(groups):
            # 创建zip文件
            zip_file_name = f"group_{i+1}.zip"
            zip_file_path = os.path.join(upload_dir, zip_file_name)
            with zipfile.ZipFile(zip_file_path, "w") as zip_file:
                # 循环添加每个文件
                for file_name in group:
                    file_path = os.path.join(upload_dir, file_name)
                    zip_file.write(file_path, arcname=file_name)

            # 生成下载链接
            download_link = downloadlink + "/filesdownload/" + zip_file_name + "/" + period
            download_links.append(download_link)

        try:
            RandomFileId.objects.create(team_one=download_links[0], team_two=download_links[1], team_three=download_links[2],
                                        team_four=download_links[3], team_five=download_links[4], team_six=download_links[5], period_id=period)

            print(f"生成下载链接成功：{download_links}")
            return Response(data={'msg': '处理成功', 'code': 200}, status=200)
        except Exception as e:
            print(e)
            return Response(data={'msg': '处理失败，请检查period', 'code': 400}, status=400)



class MergeFile(APIView):
    """
    合并各组长已上传的评分文件
    """
    authentication_classes = []
    permission_classes = []


    queryset = JudgeFiles.objects.filter()
    serializer_class = serializers.JudgeFileSerializer

    def post(self, request):
        req = request.data
        period = (req['period'])
        name = '评分汇总.xlsx'
        filepath = "upload/judge/exam" + str(period) + "/"
        file = "upload/judge/exam" + str(period) + "/" + name
        try:
            query = JudgeFiles.objects.filter(period=period).values('uploader')[::1]
            uploadedid = []
            uploader = 1
            for index in query:
                uploadedid.append(index.get('uploader'))
            if uploader in uploadedid:
                return Response(data={'msg': '处理失败，文件已合并', 'code': 405})
            else:
                obj = mergeexcel.MakeExcel()
                obj.startwork(filepath)
                jfobj = JudgeFiles(name=name, period=period, file=file, limit_period_id=1,uploader_id=1)
                jfobj.save()
                return Response(data={'msg': '处理成功', 'code': 200}, status=200)
        except Exception as e:
            print(e)
            return Response(data={'msg': '处理失败，请检查period', 'code': 400}, status=400)

class CaptainPersion(APIView):
    #为请求的组长下发随机好的组员
    """
    根据组长id下发全部判分组员信息，每期执行一次
    """
    authentication_classes = []
    permission_classes = []
    def get(self, request):
        req = request.GET
        period = (req['period'])
        captain_name = (req['username'])
        qs = RandomFileId.objects.filter(period_id=period).first()  # GenericAPIView提供的等同于self.queryset
        ser = serializers.RandomFileIdSerializer(instance=qs, many=False)
        # print(qs)
        if captain_name in  captainName and qs != None:
            return Response(data={'people':ser.data,'msg': '处理成功', 'code': 200}, status=200)
        elif qs == None:
            return Response(data={'msg': '本期暂未生成提交文件', 'code': 404}, status=200)
        else:
            return Response(data={'msg': '你没有权限，请求失败', 'code': 401}, status=200)

class CaptainUploadList(APIView):
    """
    生成组长上传答题文件下载链接返回
    """
    authentication_classes = []
    permission_classes = []
    def get(self, request):
        req = request.GET
        period = (req['period'])
        print(period)
        if not period:
            return Response(data={'msg': '处理失败，没有period参数', 'code': 401}, status=200)

        # captain_name = (req['username'])
        qs = CaptainFiles.objects.filter(period=period).values('id')
        for value in qs:   #从组长上传表查询对应期数的id，开始循环。
            id = value['id']
            address = f"{downloadlink}/captainfilesdownload/{id}/download" #赋值当前id的下载链接
            p = CaptainFiles.objects.get(id=id)
            print(p.downloadfile)
            if not p.downloadfile:   #如果当前downloadfile字段无值，将链接传入并保存到数据库中
                p.downloadfile = address
                p.save()
        query = CaptainFiles.objects.filter(period=period).all()
        ser = serializers.CaptainFileSerializer(instance=query, many=True)
        return Response(data={'data': ser.data, 'msg': '处理成功', 'code': 200}, status=200)

class CaptainJudgeUploadList(APIView):
    """
    生成组长上传评分文件下载链接返回
    """
    authentication_classes = []
    permission_classes = []
    def get(self, request):
        req = request.GET
        period = (req['period'])
        print(period)
        if not period:
            return Response(data={'msg': '处理失败，没有period参数', 'code': 401}, status=200)

        # captain_name = (req['username'])
        qs = JudgeFiles.objects.filter(period=period).values('id')
        for value in qs:
            id = value['id']
            address = f"{downloadlink}/judgefiledownload/{id}/download"
            p = JudgeFiles.objects.get(id=id)
            # print(p.downloadfile)
            if not p.downloadfile:
                p.downloadfile = address
                p.save()
        query = JudgeFiles.objects.filter(period=period).all()
        ser = serializers.JudgeFileSerializer(instance=query, many=True)
        return Response(data={'data': ser.data, 'msg': '处理成功', 'code': 200}, status=200)

class PersonalGradeUploadList(APIView):
    """
    生成个人成绩文件下载链接返回
    """
    authentication_classes = []
    permission_classes = []
    def get(self, request):
        req = request.GET
        period = (req['period'])
        # uploader = (req['uploader'])
        # print(period)
        if not period:
            return Response(data={'msg': '处理失败，没有period参数', 'code': 401}, status=200)
        # if not uploader:
            # return Response(data={'msg': '处理失败，没有uploader参数', 'code': 401}, status=200)

        # captain_name = (req['username'])
        try:
            qs = PersonalGradeFiles.objects.filter(period=period).values('id')
            print(qs)
            for value in qs:
                id = value['id']
                address = f"{downloadlink}/personalgradefiledownload/{id}/download"
                p = PersonalGradeFiles.objects.get(id=id)
                # print(p.uploader)
                # print(address)
                # print(p)
                # print('------')
                file_obj = Files.objects.get(period=period, uploader=p.uploader)
                if not file_obj.downloadfile:
                    file_obj.downloadfile = address
                    file_obj.save()
            query = Files.objects.filter(period=period).all()
            ser = serializers.FileSerializer(instance=query, many=True)
            return Response(data={'data': ser.data, 'msg': '处理成功', 'code': 200}, status=200)
        except:
            return Response(data={'msg': '未找到当期评分', 'code': 404}, status=200)

class ExamUploadList(APIView):
    """
    生成考题文件下载链接返回
    """
    authentication_classes = []
    permission_classes = []
    def get(self, request):
        req = request.GET
        period = (req['period'])
        if not period:
            return Response(data={'msg': '处理失败，没有period参数', 'code': 401}, status=200)

        # captain_name = (req['username'])
        qs = ExamFiles.objects.filter(period=period).values('id')
        for value in qs:
            id = value['id']
            address = f"{downloadlink}/examfilesdownload/{id}/download"
            p = ExamFiles.objects.get(id=id)
            # print(p.downloadfile)
            if not p.downloadfile:
                p.downloadfile = address
                p.save()
        query = ExamFiles.objects.filter(period=period).all()
        ser = serializers.ExamFileSerializer(instance=query, many=True)
        # print(ser.data)
        if query == None:
            return Response(data={'msg': '处理失败，未发现考核题文件', 'code': 404}, status=200)
        return Response(data={'data':ser.data, 'msg': '处理成功', 'code': 200}, status=200)

class UploadedEmployee(APIView):
    """
    查询本期已上传答题人员
    """
    authentication_classes = []
    permission_classes = []
    def get(self, request):
        req = request.GET
        period = (req['period'])
        query = Files.objects.filter(period=period).all()
        ser = serializers.FilePersonSerializer(instance=query, many=True)
        print(ser.data)
        return Response(data={'data':ser.data, 'msg': '处理成功', 'code': 200}, status=200)

import openpyxl
from django.apps import apps
class InsertDatabase(APIView):

    authentication_classes = []
    permission_classes = []
    def post(self, request):
        # 获取前端请求插入数据库的期数
        body_unicode = request.body.decode('utf-8')
        body = json.loads(body_unicode)
        period = body.get('period', '')
        period = str(period)

        # 根据相应的期数获取需要写入的模型类
        score_model_name = 'Score' + period
        linux_model_name = 'LinuxScore' + period
        network_model_name = 'NetworkScore' + period
        office_model_name = 'OfficeScore' + period
        try:
            ScoreModel = apps.get_model('hsoftskill', score_model_name)
            LinuxScoreModel = apps.get_model('hsoftskill', linux_model_name)
            NetworkScoreModel = apps.get_model('hsoftskill', network_model_name)
            OfficeScoreModel = apps.get_model('hsoftskill', office_model_name)
        except LookupError:
            OfficeScoreModel = None

        if OfficeScoreModel is None:
            print("本期无文档题")
        # 读取Excel文件
        wb = openpyxl.load_workbook(filerootpath +'/judge/exam' + period + '/评分汇总.xlsx')

        try:
            total_sheet = wb['总成绩']
            network_sheet = wb['数通题']
            linux_sheet = wb['服务器题']
            office_sheet = wb['文档题']
        except LookupError:
            office_sheet = None
        if office_sheet is None:
            pass

        # 读取数据
        if period == '3':
            try:
                if not (ScoreModel.objects.exists() or LinuxScoreModel.objects.exists() or NetworkScoreModel.objects.exists()
                        or OfficeScoreModel.objects.exists()):

                    for row in total_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, office_score, network_score, linux_score = row
                        score = ScoreModel(name=name, office_score=office_score, network_score=network_score,
                                           linux_score=linux_score)
                        score.save()

                    for row in linux_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, item7, item8, item9, item10, total = row
                        linux_score = LinuxScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                                      , item4=item4, item5=item5, item6=item6, item7=item7,
                                                      item8=item8, item9=item9,  item10=item10, total=total)
                        linux_score.save()

                    for row in network_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, item7, item8, total = row
                        network_score = NetworkScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                , item4=item4, item5=item5, item6=item6, item7=item7, item8=item8, total=total)
                        network_score.save()

                    for row in office_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, item7, item8, total = row
                        office_score = OfficeScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                        , item4=item4, item5=item5, item6=item6, item7=item7, item8=item8, total=total)
                        office_score.save()

                    return Response(data={'msg': '数据添加成功', 'code': 200}, status=200)
                else:
                    return Response(data={'msg': '添加失败，已经插入过数据', 'code': 403}, status=200)
            # do something
            except NameError:
                # OfficeScoreModel doesn't exist, continue with other logic
                pass
        elif period == '4':
            try:
                if not (ScoreModel.objects.exists() or LinuxScoreModel.objects.exists() or NetworkScoreModel.objects.exists()
                        or OfficeScoreModel.objects.exists()):

                    for row in total_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, office_score, network_score, linux_score = row
                        score = ScoreModel(name=name, office_score=office_score, network_score=network_score,
                                           linux_score=linux_score)
                        score.save()

                    for row in linux_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, item7, item8, item9,total = row
                        linux_score = LinuxScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                                      , item4=item4, item5=item5, item6=item6, item7=item7,
                                                      item8=item8, item9=item9, total=total)
                        linux_score.save()

                    for row in network_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, item7, item8, item9, item10, item11, item12, total = row
                        network_score = NetworkScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                , item4=item4, item5=item5, item6=item6, item7=item7, item8=item8
                                                          , item9=item9, item10=item10, item11=item11, item12=item12, total=total)
                        network_score.save()

                    for row in office_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, item7, item8, item9, total = row
                        office_score = OfficeScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                                        , item4=item4, item5=item5, item6=item6, item7=item7, item8=item8,
                                                        item9=item9, total=total)
                        office_score.save()

                    return Response(data={'msg': '数据添加成功', 'code': 200}, status=200)
                else:
                    return Response(data={'msg': '添加失败，已经插入过数据', 'code': 403}, status=200)
            # do something
            except NameError:
                # OfficeScoreModel doesn't exist, continue with other logic
                pass
        elif period == '5':
            try:
                if not (ScoreModel.objects.exists() or LinuxScoreModel.objects.exists() or NetworkScoreModel.objects.exists()
                        or OfficeScoreModel.objects.exists()):

                    for row in total_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, office_score, network_score, linux_score = row
                        score = ScoreModel(name=name, office_score=office_score, network_score=network_score, linux_score=linux_score)
                        score.save()

                    for row in linux_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, item7, total = row
                        linux_score = LinuxScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                                      , item4=item4, item5=item5, item6=item6, item7=item7, total=total)
                        linux_score.save()

                    for row in network_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, item7, total = row
                        network_score = NetworkScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                                          , item4=item4, item5=item5, item6=item6, item7=item7, total=total)
                        network_score.save()

                    for row in office_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, item7, item8,total = row
                        office_score = OfficeScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                            , item4=item4, item5=item5, item6=item6, item7=item7, item8=item8, total=total)
                        office_score.save()

                    return Response(data={'msg': '数据添加成功', 'code': 200}, status=200)
                else:
                    return Response(data={'msg': '添加失败，已经插入过数据', 'code': 403}, status=200)
            # do something
            except NameError:
                # OfficeScoreModel doesn't exist, continue with other logic
                pass
        elif period == '6':
            try:
                if not (ScoreModel.objects.exists() or LinuxScoreModel.objects.exists() or NetworkScoreModel.objects.exists()):

                    for row in total_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, network_score, linux_score = row
                        score = ScoreModel(name=name, network_score=network_score, linux_score=linux_score)
                        score.save()

                    for row in linux_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, total = row
                        linux_score = LinuxScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                                          , item4=item4, total=total)
                        linux_score.save()

                    for row in network_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, item7, total = row
                        network_score = NetworkScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                                          , item4=item4, item5=item5, item6=item6, item7=item7, total=total)
                        network_score.save()

                    return Response(data={'msg': '数据添加成功', 'code': 200}, status=200)
                else:
                    return Response(data={'msg': '添加失败，已经插入过数据', 'code': 403}, status=200)
            # do something
            except NameError:
                # OfficeScoreModel doesn't exist, continue with other logic
                pass
        elif period == '7':
            try:
                if not (
                        ScoreModel.objects.exists() or LinuxScoreModel.objects.exists() or NetworkScoreModel.objects.exists()):

                    for row in total_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, network_score, linux_score = row
                        score = ScoreModel(name=name, network_score=network_score, linux_score=linux_score)
                        score.save()

                    for row in linux_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, total = row
                        linux_score = LinuxScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                                          , item4=item4, item5=item5, item6=item6, total=total)
                        linux_score.save()

                    for row in network_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, total = row
                        network_score = NetworkScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                                          , item4=item4, item5=item5, item6=item6, total=total)
                        network_score.save()

                    return Response(data={'msg': '数据添加成功', 'code': 200}, status=200)
                else:
                    return Response(data={'msg': '添加失败，已经插入过数据', 'code': 403}, status=200)
            # do something
            except NameError:
                # OfficeScoreModel doesn't exist, continue with other logic
                pass

        elif period == '8':
            try:
                if not (
                        ScoreModel.objects.exists() or LinuxScoreModel.objects.exists() or NetworkScoreModel.objects.exists()):

                    for row in total_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, network_score, linux_score = row
                        score = ScoreModel(name=name, network_score=network_score, linux_score=linux_score)
                        score.save()

                    for row in linux_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, total = row
                        linux_score = LinuxScoreModel(name=name, item1=item1, item2=item2, item3=item3, item4=item4,
                                                      total=total)
                        linux_score.save()

                    for row in network_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, total = row
                        network_score = NetworkScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                                          , item4=item4, item5=item5, total=total)
                        network_score.save()

                    return Response(data={'msg': '数据添加成功', 'code': 200}, status=200)
                else:
                    return Response(data={'msg': '添加失败，已经插入过数据', 'code': 403}, status=200)
            # do something
            except NameError:
                # OfficeScoreModel doesn't exist, continue with other logic
                pass

        elif period == '9':
            try:
                if not (ScoreModel.objects.exists() or LinuxScoreModel.objects.exists() or NetworkScoreModel.objects.exists()):
                    for row in total_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, network_score, linux_score = row
                        score = ScoreModel(name=name, network_score=network_score, linux_score=linux_score)
                        score.save()

                    for row in linux_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, total = row
                        linux_score = LinuxScoreModel(name=name, item1=item1, item2=item2, item3=item3, item4=item4,
                                                      total=total)
                        linux_score.save()

                    for row in network_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, total = row
                        network_score = NetworkScoreModel(name=name, item1=item1, item2=item2, item3=item3
                                                          , item4=item4, item5=item5, item6=item6, total=total)
                        network_score.save()

                    return Response(data={'msg': '数据添加成功', 'code': 200}, status=200)
                else:
                    return Response(data={'msg': '添加失败，已经插入过数据', 'code': 403}, status=200)
            # do something
            except NameError:
                # OfficeScoreModel doesn't exist, continue with other logic
                pass

        elif period == '10':
            try:
                if not (ScoreModel.objects.exists() or LinuxScoreModel.objects.exists()):
                    for row in total_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, network_score, linux_score = row
                        score = ScoreModel(name=name, network_score=network_score, linux_score=linux_score)
                        score.save()

                    for row in linux_sheet.iter_rows(min_row=3, min_col=2, values_only=True):
                        name, item1, item2, item3, item4, item5, item6, item7, item8, total = row
                        linux_score = LinuxScoreModel(name=name, item1=item1, item2=item2, item3=item3, item4=item4,
                                                      item5=item5, item6=item6,  item7=item7, item8=item8, total=total)
                        linux_score.save()

                    return Response(data={'msg': '数据添加成功', 'code': 200}, status=200)
                else:
                    return Response(data={'msg': '添加失败，已经插入过数据', 'code': 403}, status=200)
            except:
                return Response(data={'msg': '未找到当前期数成绩信息', 'code': 404}, status=200)

        else:
            return Response(data={'msg': '未找到当前期数成绩信息', 'code': 404}, status=200)

class GetAllPeriod(APIView):
    '''
    获取所有考核期数
    '''
    authentication_classes = []
    permission_classes = []
    def get(self, request):
        count = LimitFile.objects.count()
        periods = [{'id': i} for i in range(1, count + 1)]
        return Response(periods)

class GetScore(APIView):
    '''
    获取指定期数成绩信息
    '''
    from django.apps import apps
    from django.http import JsonResponse

    def get(self, request):
        period = request.GET['period']

        score_model_name = 'Score' + period
        linux_model_name = 'LinuxScore' + period
        network_model_name = 'NetworkScore' + period
        office_model_name = 'OfficeScore' + period
        try:
            ScoreModel = apps.get_model('hsoftskill', score_model_name)
            LinuxScoreModel = apps.get_model('hsoftskill', linux_model_name)
            NetworkScoreModel = apps.get_model('hsoftskill', network_model_name)
            OfficeScoreModel = apps.get_model('hsoftskill', office_model_name)
        except:
            OfficeScoreModel = None

        if OfficeScoreModel is None:
            print("本期无文档题")
        try:
            scores = ScoreModel.objects.all().values()
        except:
            scores = []
        try:
            linux_scores = LinuxScoreModel.objects.all().values()
        except:
            linux_scores = []
        try:
            network_scores = NetworkScoreModel.objects.all().values()
        except:
            network_scores = []
        try:
            office_scores = OfficeScoreModel.objects.all().values()
        except:
            office_scores = []
        try:
            fields = [
                {'key': field.name, 'label': field.verbose_name} for field in ScoreModel._meta.fields if
                            field.name not in ['period']
                        ]
        except:
            fields = []
        try:
            linux_fields = [{'key': field.name, 'label': field.verbose_name} for field in LinuxScoreModel._meta.fields if
                            field.name not in ['period']
                            ]
        except:
            linux_fields = []
        try:
            network_fields = [
                {'key': field.name, 'label': field.verbose_name} for field in NetworkScoreModel._meta.fields if
                            field.name not in ['period']
                         ]
        except:
            network_fields = []
        try:
            office_fields = [
                {'key': field.name, 'label': field.verbose_name} for field in OfficeScoreModel._meta.fields if
                            field.name not in ['period']
                ]
        except:
            office_fields = []
        try:
            return JsonResponse(
                {'scores': list(scores), 'linux_scores': list(linux_scores), 'network_scores': list(network_scores),
                 'office_scores': list(office_scores),
                 'fields': fields, 'linux_fields': linux_fields, 'network_fields': network_fields, 'office_fields': office_fields}
            )
        except:
            return JsonResponse({'msg': '未找到当前期数成绩信息', 'code': 404})